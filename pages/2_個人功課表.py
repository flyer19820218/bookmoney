import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.page import PageMargins
import re
import io
import requests
import os

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

st.set_page_config(page_title="個人功課表產生器", layout="centered", page_icon="🗓️")

FONT_NAME = "NotoSansTC-Regular.ttf"
FONT_URL = "https://raw.githubusercontent.com/google/fonts/main/ofl/notosanstc/NotoSansTC%5Bwght%5D.ttf"

@st.cache_resource
def init_fonts():
    """確保 PDF 可正確顯示繁體中文。"""
    if not os.path.exists(FONT_NAME):
        try:
            with st.spinner("正在準備 PDF 中文字型..."):
                response = requests.get(
                    FONT_URL,
                    headers={"User-Agent": "Mozilla/5.0"},
                    timeout=60,
                )
                response.raise_for_status()
                with open(FONT_NAME, "wb") as font_file:
                    font_file.write(response.content)
        except Exception as error:
            st.error(f"中文字型下載失敗：{error}")
            return False
    try:
        pdfmetrics.registerFont(TTFont("CustomFont", FONT_NAME))
        return True
    except Exception as error:
        st.error(f"中文字型註冊失敗：{error}")
        return False

HAS_FONT = init_fonts()

st.title("🗓️ 個人功課表")

st.subheader("個人功課表產生器")
st.info("""
系統先以老師辨識分組節次，再依數理、英文與資優名單安排跑班。
**輸出只保留正式課表上的科目名稱**，不會把內部判斷的真實課程寫出去。
自動判斷後一定會先顯示「節次規則」及「學生分組」供老師校正，確認後才產生課表。
""")

if pdfplumber is None:
    st.error("目前環境缺少 pdfplumber。請在 requirements.txt 加入 `pdfplumber` 後重新部署。")
    st.stop()

DAY_NAMES = ["星期一", "星期二", "星期三", "星期四", "星期五"]
PERIOD_LABELS = {
    0: "早修", 1: "第1節", 2: "第2節", 3: "第3節", 4: "第4節",
    5: "第5節", 6: "第6節", 7: "第7節", 8: "第8節", 9: "第9節",
}
PERIOD_NUMBERS = {v: k for k, v in PERIOD_LABELS.items()}
RULE_NORMAL = "不分組"
RULE_STEM = "數理分組"
RULE_ENGLISH = "英文分組"
RULE_OPTIONS = [RULE_NORMAL, RULE_STEM, RULE_ENGLISH]

# 本次901已確認的三位數理資優生；之後仍可在校正表直接修改。
VERIFIED_STEM_GIFTED = {"陳昱學", "梁容嘉", "洪采漩"}

def clean_pdf_lines(value):
    if value is None:
        return []
    result = []
    for line in str(value).splitlines():
        line = re.sub(r"\s+", "", line).strip()
        if line:
            result.append(line)
    return result

def parse_course_cell(value):
    lines = clean_pdf_lines(value)
    return {
        "subject": lines[0] if lines else "",
        "teacher": lines[1] if len(lines) >= 2 else "",
        "raw": "\n".join(lines),
    }

@st.cache_data(show_spinner=False)
def parse_timetable_pdf(pdf_bytes):
    schedules = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            # 部分學校PDF會把 901 抽成「9 01」，因此先容許班號中含空白。
            class_match = re.search(r"班級\s*[：:]\s*([0-9\s]{3,6})", page_text)
            if not class_match:
                continue
            class_digits = re.sub(r"\s+", "", class_match.group(1))
            if len(class_digits) < 3:
                continue
            class_no = int(class_digits[:3])
            tables = page.extract_tables() or []
            if not tables:
                continue
            table = max(tables, key=lambda t: len(t) if t else 0)
            class_schedule = {}

            # pdfplumber表格：第1列標題、第2列早修、第3～11列為第1～9節。
            for row_index, row in enumerate(table):
                if row_index == 1:
                    period_no = 0
                elif 2 <= row_index <= 10:
                    period_no = row_index - 1
                else:
                    continue

                row = list(row or [])
                while len(row) < 8:
                    row.append("")
                for day_index in range(5):
                    class_schedule[(period_no, day_index)] = parse_course_cell(row[3 + day_index])

            schedules[class_no] = class_schedule
    return schedules

@st.cache_data(show_spinner=False)
def parse_group_pdf(pdf_bytes, group_kind):
    rows = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        if not pdf.pages:
            return rows
        page = pdf.pages[0]
        page_text = page.extract_text() or ""
        tables = page.extract_tables() or []

        if group_kind == "數理":
            markers = re.findall(r"(\d{3})\s*數理\s*([A-Za-z])", page_text)
        else:
            markers = re.findall(r"(\d{3})\s*英\s*([A-Za-z])", page_text)

        if len(markers) < len(tables):
            raise ValueError(f"無法從{group_kind}分組PDF辨識每一欄的目的班級。")

        for table_index, table in enumerate(tables):
            destination_class = int(markers[table_index][0])
            group_letter = markers[table_index][1].upper()
            group_label = f"{destination_class}{group_kind}{group_letter}"
            grade_base = (destination_class // 100) * 100

            for source_row in (table or [])[1:]:
                source_row = list(source_row or [])
                while len(source_row) < 5:
                    source_row.append("")
                original_class = str(source_row[1] or "").strip()
                seat = str(source_row[2] or "").strip()
                student_id = str(source_row[3] or "").strip()
                name = re.sub(r"\s+", "", str(source_row[4] or ""))

                if not name or not original_class:
                    continue
                try:
                    home_class = grade_base + int(float(original_class))
                    seat_no = int(float(seat))
                except (TypeError, ValueError):
                    continue

                student_key = student_id if student_id else f"{home_class}-{seat_no}-{name}"
                rows.append({
                    "學生識別": student_key,
                    "學號": student_id,
                    "姓名": name,
                    "原班": home_class,
                    "座號": seat_no,
                    "分組種類": group_kind,
                    "分組代號": group_label,
                    "目的班級": destination_class,
                })
    return rows

def build_student_group_table(math_rows, english_rows):
    students = {}
    for group_rows in [math_rows, english_rows]:
        for item in group_rows:
            key = item["學生識別"]
            if key not in students:
                students[key] = {
                    "學生識別": key,
                    "原班": item["原班"],
                    "座號": item["座號"],
                    "學號": item["學號"],
                    "姓名": item["姓名"],
                    "數理去向": "未設定",
                    "英文去向": "未設定",
                    "資優類別": "無",
                }
            if item["分組種類"] == "數理":
                students[key]["數理去向"] = item["分組代號"]
            else:
                students[key]["英文去向"] = item["分組代號"]

    result = pd.DataFrame(students.values())
    if result.empty:
        return result
    for idx, student in result.iterrows():
        if student["姓名"] in VERIFIED_STEM_GIFTED:
            result.at[idx, "資優類別"] = "數理資優"
    return result.sort_values(["原班", "座號"]).reset_index(drop=True)

def teacher_seed_pools(schedules, class_numbers):
    pools = {}
    stem_words = ["數學", "自然科學", "理化", "生物", "地科", "STEM", "數學專題"]
    for class_no in class_numbers:
        stem_teachers = set()
        english_teachers = set()
        for cell in schedules.get(class_no, {}).values():
            subject = str(cell.get("subject", ""))
            teacher = str(cell.get("teacher", ""))
            if not teacher:
                continue
            if any(word in subject for word in stem_words):
                stem_teachers.add(teacher)
            if "英語文" in subject:
                english_teachers.add(teacher)
        pools[class_no] = {"stem": stem_teachers, "english": english_teachers}
    return pools

def infer_slot_rules(schedules, cluster_classes, gifted_class):
    all_classes = list(cluster_classes)
    if gifted_class:
        all_classes.append(gifted_class)
    pools = teacher_seed_pools(schedules, all_classes)
    rules = {}

    for period_no in range(10):
        for day_index in range(5):
            english_hits = 0
            stem_hits = 0
            for class_no in cluster_classes:
                teacher = schedules.get(class_no, {}).get((period_no, day_index), {}).get("teacher", "")
                if teacher and teacher in pools.get(class_no, {}).get("english", set()):
                    english_hits += 1
                if teacher and teacher in pools.get(class_no, {}).get("stem", set()):
                    stem_hits += 1

            gifted_teacher = ""
            gifted_is_stem = False
            if gifted_class:
                gifted_teacher = schedules.get(gifted_class, {}).get((period_no, day_index), {}).get("teacher", "")
                gifted_is_stem = bool(
                    gifted_teacher and
                    gifted_teacher in pools.get(gifted_class, {}).get("stem", set())
                )

            if english_hits == len(cluster_classes) and english_hits > 0:
                rules[(period_no, day_index)] = RULE_ENGLISH
            elif gifted_is_stem and stem_hits >= max(2, len(cluster_classes) - 1):
                # 914同節有數理資優課，且至少兩個跑班教室出現數理老師。
                # 可涵蓋「科目名稱或其中一位老師被偽裝」的節次。
                rules[(period_no, day_index)] = RULE_STEM
            else:
                rules[(period_no, day_index)] = RULE_NORMAL
    return rules

def compact_cell(cell):
    if not cell or not cell.get("raw"):
        return ""
    return "／".join(clean_pdf_lines(cell.get("raw", "")))

def build_rule_editor_table(schedules, cluster_classes, gifted_class, inferred_rules):
    rows = []
    visible_classes = list(cluster_classes)
    if gifted_class:
        visible_classes.append(gifted_class)
    for period_no in range(10):
        for day_index, day_name in enumerate(DAY_NAMES):
            row = {
                "星期": day_name,
                "節次": PERIOD_LABELS[period_no],
                "課程類型": inferred_rules[(period_no, day_index)],
            }
            for class_no in visible_classes:
                row[str(class_no)] = compact_cell(
                    schedules.get(class_no, {}).get((period_no, day_index), {})
                )
            rows.append(row)
    return pd.DataFrame(rows)

def route_class_from_label(route_label):
    match = re.match(r"\s*(\d{3})", str(route_label))
    return int(match.group(1)) if match else None

def make_personal_grid(student, schedules, rule_map, gifted_class):
    home_class = int(student["原班"])
    gifted_type = str(student.get("資優類別", "無"))
    result = {}
    errors = []

    for period_no in range(10):
        for day_index in range(5):
            rule = rule_map.get((period_no, day_index), RULE_NORMAL)
            destination_class = home_class
            route_label = "原班"
            cell_kind = "normal"

            if rule == RULE_STEM:
                cell_kind = "stem"
                if "數理資優" in gifted_type:
                    destination_class = gifted_class
                    route_label = f"{gifted_class}數理資優"
                    cell_kind = "gifted"
                else:
                    route_label = str(student.get("數理去向", "未設定"))
                    destination_class = route_class_from_label(route_label)
            elif rule == RULE_ENGLISH:
                cell_kind = "english"
                if "語文資優" in gifted_type:
                    destination_class = gifted_class
                    route_label = f"{gifted_class}語文資優"
                    cell_kind = "gifted"
                else:
                    route_label = str(student.get("英文去向", "未設定"))
                    destination_class = route_class_from_label(route_label)

            if destination_class is None or destination_class not in schedules:
                errors.append(
                    f"{student['姓名']}：{DAY_NAMES[day_index]}{PERIOD_LABELS[period_no]}的去向「{route_label}」無有效課表。"
                )
                cell = {"subject": "⚠️去向未設定", "teacher": "", "raw": "⚠️去向未設定"}
                cell_kind = "warning"
            else:
                cell = schedules.get(destination_class, {}).get(
                    (period_no, day_index), {"subject": "", "teacher": "", "raw": ""}
                )

            official_lines = clean_pdf_lines(cell.get("raw", ""))
            display_lines = list(official_lines)
            if rule != RULE_NORMAL:
                # 即使目的班級與原班相同，也仍標示為跑班課。
                display_lines.append(f"【跑班：{route_label}／{destination_class}教室】")
                if not official_lines:
                    errors.append(
                        f"{student['姓名']}：{DAY_NAMES[day_index]}{PERIOD_LABELS[period_no]}的"
                        f"{destination_class}課表是空白。"
                    )
                    display_lines.insert(0, "⚠️正式課表空白")
                    cell_kind = "warning"

            result[(period_no, day_index)] = {
                "text": "\n".join(display_lines),
                "kind": cell_kind,
            }
    return result, errors

def personal_grid_dataframe(grid):
    rows = []
    for period_no in range(10):
        row = {"節次": PERIOD_LABELS[period_no]}
        for day_index, day_name in enumerate(DAY_NAMES):
            row[day_name] = grid[(period_no, day_index)]["text"]
        rows.append(row)
    return pd.DataFrame(rows).set_index("節次")

def write_personal_sheet(ws, student, grid):
    title = f"{int(student['原班'])}班 {int(student['座號'])}號 {student['姓名']} 個人功課表"
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(name="微軟正黑體", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        f"數理：{student.get('數理去向', '未設定')}　"
        f"英文：{student.get('英文去向', '未設定')}　"
        f"資優：{student.get('資優類別', '無')}"
    )
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["節次"] + DAY_NAMES
    for column_no, header in enumerate(headers, 1):
        cell = ws.cell(3, column_no, header)
        cell.font = Font(name="微軟正黑體", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fills = {
        "normal": PatternFill("solid", fgColor="FFFFFF"),
        "stem": PatternFill("solid", fgColor="DDEBF7"),
        "english": PatternFill("solid", fgColor="FCE4D6"),
        "gifted": PatternFill("solid", fgColor="E4DFEC"),
        "warning": PatternFill("solid", fgColor="FFC7CE"),
    }
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for period_no in range(10):
        excel_row = period_no + 4
        label_cell = ws.cell(excel_row, 1, PERIOD_LABELS[period_no])
        label_cell.font = Font(name="微軟正黑體", bold=True)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = border
        ws.row_dimensions[excel_row].height = 58

        for day_index in range(5):
            item = grid[(period_no, day_index)]
            cell = ws.cell(excel_row, day_index + 2, item["text"])
            cell.font = Font(name="微軟正黑體", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = fills[item["kind"]]
            cell.border = border

    for col in "BCDEF":
        ws.column_dimensions[col].width = 24
    ws.column_dimensions["A"].width = 10
    ws.freeze_panes = "B4"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4)

def generate_schedule_workbook(student_records, schedules, rule_map, gifted_class):
    wb = Workbook()
    wb.remove(wb.active)
    all_errors = []
    used_sheet_names = set()

    for student in student_records:
        grid, errors = make_personal_grid(student, schedules, rule_map, gifted_class)
        all_errors.extend(errors)
        base_name = f"{int(student['座號']):02d}{student['姓名']}"[:31]
        sheet_name = base_name
        suffix = 2
        while sheet_name in used_sheet_names:
            sheet_name = f"{base_name[:27]}_{suffix}"
            suffix += 1
        used_sheet_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        write_personal_sheet(ws, student, grid)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), all_errors

def _pdf_wrapped_lines(text, font_name, font_size, max_width):
    """依實際字寬切行，保留課表原本的換行。"""
    wrapped = []
    pdf_text = str(text or "").replace("【", "[").replace("】", "]").replace("／", "/")
    for source_line in pdf_text.splitlines() or [""]:
        current = ""
        for char in source_line:
            candidate = current + char
            if current and pdfmetrics.stringWidth(candidate, font_name, font_size) > max_width:
                wrapped.append(current)
                current = char
            else:
                current = candidate
        wrapped.append(current)
    return wrapped

def draw_personal_schedule_pdf_page(pdf, student, grid):
    """在既有 PDF 畫布加入一位學生的一頁橫式 A4 功課表。"""
    page_width, page_height = A4[1], A4[0]
    font_name = "CustomFont" if HAS_FONT else "Helvetica"

    title = f"{int(student['原班'])}班 {int(student['座號'])}號 {student['姓名']} 個人功課表"
    subtitle = (
        f"數理：{student.get('數理去向', '未設定')}　"
        f"英文：{student.get('英文去向', '未設定')}　"
        f"資優：{student.get('資優類別', '無')}"
    )
    pdf.setTitle(title)
    pdf.setFont(font_name, 16)
    pdf.drawCentredString(page_width / 2, page_height - 28, title)
    pdf.setFont(font_name, 9)
    pdf.drawCentredString(page_width / 2, page_height - 45, subtitle)

    left = 20
    bottom = 20
    table_top = page_height - 58
    table_height = table_top - bottom
    header_height = 27
    row_height = (table_height - header_height) / 10
    first_width = 58
    day_width = (page_width - 40 - first_width) / 5
    widths = [first_width] + [day_width] * 5
    headers = ["節次"] + DAY_NAMES

    fill_colors = {
        "normal": (1, 1, 1),
        "stem": (0.87, 0.92, 0.97),
        "english": (0.99, 0.89, 0.84),
        "gifted": (0.89, 0.87, 0.93),
        "warning": (1, 0.78, 0.81),
    }

    x = left
    for col_no, header in enumerate(headers):
        width = widths[col_no]
        pdf.setFillColorRGB(0.27, 0.45, 0.71)
        pdf.rect(x, table_top - header_height, width, header_height, fill=1, stroke=0)
        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont(font_name, 9)
        pdf.drawCentredString(x + width / 2, table_top - 18, header)
        x += width

    for period_no in range(10):
        y_top = table_top - header_height - period_no * row_height
        x = left
        pdf.setFillColorRGB(0.95, 0.95, 0.95)
        pdf.rect(x, y_top - row_height, first_width, row_height, fill=1, stroke=0)
        pdf.setFillColorRGB(0, 0, 0)
        pdf.setFont(font_name, 8.5)
        pdf.drawCentredString(x + first_width / 2, y_top - row_height / 2 - 3, PERIOD_LABELS[period_no])
        x += first_width

        for day_index in range(5):
            item = grid[(period_no, day_index)]
            rgb = fill_colors.get(item["kind"], fill_colors["normal"])
            pdf.setFillColorRGB(*rgb)
            pdf.rect(x, y_top - row_height, day_width, row_height, fill=1, stroke=0)
            pdf.setFillColorRGB(0, 0, 0)
            font_size = 7.2
            lines = _pdf_wrapped_lines(item["text"], font_name, font_size, day_width - 8)
            if len(lines) > 5:
                font_size = 6.4
                lines = _pdf_wrapped_lines(item["text"], font_name, font_size, day_width - 8)
            lines = lines[:6]
            line_height = font_size + 1.2
            text_y = y_top - (row_height - len(lines) * line_height) / 2 - font_size
            pdf.setFont(font_name, font_size)
            for line in lines:
                pdf.drawCentredString(x + day_width / 2, text_y, line)
                text_y -= line_height
            x += day_width

    pdf.setStrokeColorRGB(0.5, 0.5, 0.5)
    x = left
    for width in widths:
        pdf.rect(x, bottom, width, table_height, fill=0, stroke=1)
        x += width
    for row_no in range(11):
        y = table_top - header_height - row_no * row_height
        pdf.line(left, y, page_width - 20, y)
    pdf.line(left, table_top, page_width - 20, table_top)

    pdf.showPage()

def generate_personal_schedule_pdf(student, grid):
    """製作單一學生、一頁橫式 A4 的個人功課表 PDF。"""
    output = io.BytesIO()
    page_width, page_height = A4[1], A4[0]
    pdf = canvas.Canvas(output, pagesize=(page_width, page_height))
    draw_personal_schedule_pdf_page(pdf, student, grid)
    pdf.save()
    return output.getvalue()

def generate_all_schedule_pdf(student_records, schedules, rule_map, gifted_class):
    """製作全班共用的多頁 PDF，每位學生固定一頁。"""
    output = io.BytesIO()
    all_errors = []
    page_width, page_height = A4[1], A4[0]
    pdf = canvas.Canvas(output, pagesize=(page_width, page_height))
    for student in student_records:
        grid, errors = make_personal_grid(student, schedules, rule_map, gifted_class)
        all_errors.extend(errors)
        draw_personal_schedule_pdf_page(pdf, student, grid)
    pdf.save()
    return output.getvalue(), all_errors

st.markdown("#### 步驟一：上傳正式課表與分組名單")
col_t1, col_t2, col_t3 = st.columns(3)
with col_t1:
    timetable_file = st.file_uploader(
        "正式課表 PDF（同一跑班班群＋資優班）", type=["pdf"], key="personal_timetable_pdf"
    )
with col_t2:
    timetable_math_file = st.file_uploader(
        "數理分組名單 PDF", type=["pdf"], key="personal_math_pdf"
    )
with col_t3:
    timetable_english_file = st.file_uploader(
        "英文分組名單 PDF", type=["pdf"], key="personal_english_pdf"
    )

if timetable_file and timetable_math_file and timetable_english_file:
    try:
        schedules = parse_timetable_pdf(timetable_file.getvalue())
        math_rows = parse_group_pdf(timetable_math_file.getvalue(), "數理")
        english_rows = parse_group_pdf(timetable_english_file.getvalue(), "英")
        students_df = build_student_group_table(math_rows, english_rows)

        if len(schedules) < 3:
            raise ValueError("正式課表PDF至少需要包含三個跑班班級。")
        if students_df.empty:
            raise ValueError("分組PDF沒有讀到學生資料。")

        schedule_classes = sorted(schedules.keys())
        default_gifted = 914 if 914 in schedule_classes else schedule_classes[-1]
        regular_defaults = [c for c in schedule_classes if c != default_gifted][:3]

        st.success(
            f"已讀取 {len(schedules)} 份正式課表、{len(students_df)} 位學生的分組資料。"
        )

        col_cfg1, col_cfg2 = st.columns(2)
        with col_cfg1:
            cluster_classes = st.multiselect(
                "本次跑班班群（請選三班）",
                options=schedule_classes,
                default=regular_defaults,
                key="personal_cluster_classes",
            )
        with col_cfg2:
            gifted_class = st.selectbox(
                "資優班課表",
                options=schedule_classes,
                index=schedule_classes.index(default_gifted),
                key="personal_gifted_class",
            )

        if len(cluster_classes) != 3:
            st.warning("請選擇剛好三個跑班班級。")
            st.stop()
        if gifted_class in cluster_classes:
            st.warning("資優班不能同時列入三個一般跑班班級。")
            st.stop()

        inferred_rules = infer_slot_rules(schedules, cluster_classes, gifted_class)
        rule_df = build_rule_editor_table(schedules, cluster_classes, gifted_class, inferred_rules)
        auto_grouped = rule_df[rule_df["課程類型"] != RULE_NORMAL]

        st.markdown("#### 步驟二：確認哪些節次需要跑班")
        st.caption(
            "自動判斷只看老師與三班／資優班同節課表；科目名稱僅供老師核對，不作為最終依據。"
        )
        count_stem = int((rule_df["課程類型"] == RULE_STEM).sum())
        count_english = int((rule_df["課程類型"] == RULE_ENGLISH).sum())
        st.write(f"目前判定：數理跑班 **{count_stem}節**、英文跑班 **{count_english}節**。")
        st.dataframe(auto_grouped, use_container_width=True, hide_index=True)

        with st.expander("老師校正：分組／不分組節次", expanded=False):
            editable_rule_columns = ["星期", "節次", "課程類型"] + [str(c) for c in cluster_classes] + [str(gifted_class)]
            edited_rule_df = st.data_editor(
                rule_df[editable_rule_columns],
                use_container_width=True,
                hide_index=True,
                height=520,
                disabled=["星期", "節次"] + [str(c) for c in cluster_classes] + [str(gifted_class)],
                column_config={
                    "課程類型": st.column_config.SelectboxColumn(
                        "課程類型", options=RULE_OPTIONS, required=True
                    )
                },
                key="personal_rule_editor",
            )

        rule_map = {}
        for _, row in edited_rule_df.iterrows():
            day_index = DAY_NAMES.index(row["星期"])
            period_no = PERIOD_NUMBERS[row["節次"]]
            rule_map[(period_no, day_index)] = row["課程類型"]

        available_home_classes = sorted(
            c for c in students_df["原班"].dropna().astype(int).unique().tolist()
            if c in cluster_classes
        )
        if not available_home_classes:
            available_home_classes = sorted(students_df["原班"].dropna().astype(int).unique().tolist())

        st.markdown("#### 步驟三：老師校正學生分組與資優身分")
        target_class = st.selectbox(
            "要產生哪一班的個人課表？",
            options=available_home_classes,
            index=available_home_classes.index(901) if 901 in available_home_classes else 0,
            key="personal_target_class",
        )

        class_students = students_df[students_df["原班"] == target_class].copy().reset_index(drop=True)
        math_options = ["未設定"] + sorted({r["分組代號"] for r in math_rows})
        english_options = ["未設定"] + sorted({r["分組代號"] for r in english_rows})

        edited_students = st.data_editor(
            class_students[["原班", "座號", "學號", "姓名", "數理去向", "英文去向", "資優類別"]],
            use_container_width=True,
            hide_index=True,
            disabled=["原班", "座號", "學號", "姓名"],
            column_config={
                "數理去向": st.column_config.SelectboxColumn(
                    "數理去向", options=math_options, required=True
                ),
                "英文去向": st.column_config.SelectboxColumn(
                    "英文去向", options=english_options, required=True
                ),
                "資優類別": st.column_config.SelectboxColumn(
                    "資優類別",
                    options=["無", "數理資優", "語文資優", "數理／語文資優"],
                    required=True,
                ),
            },
            key=f"personal_student_editor_{target_class}",
        )

        missing_math = edited_students[edited_students["數理去向"] == "未設定"]
        missing_english = edited_students[edited_students["英文去向"] == "未設定"]
        if not missing_math.empty or not missing_english.empty:
            st.error(
                f"仍有 {len(missing_math)} 位學生未設定數理去向、"
                f"{len(missing_english)} 位學生未設定英文去向；請先校正。"
            )
            st.stop()

        st.markdown("#### 步驟四：預覽與下載個人功課表")
        student_labels = {
            f"{int(row['座號']):02d}號 {row['姓名']}": idx
            for idx, row in edited_students.iterrows()
        }
        selected_label = st.selectbox("選擇學生預覽", options=list(student_labels.keys()))
        selected_student = edited_students.loc[student_labels[selected_label]].to_dict()
        selected_grid, selected_errors = make_personal_grid(
            selected_student, schedules, rule_map, gifted_class
        )
        st.dataframe(
            personal_grid_dataframe(selected_grid),
            use_container_width=True,
            height=470,
        )

        if selected_errors:
            with st.expander(f"此學生有 {len(selected_errors)} 個需要確認的地方", expanded=True):
                for message in selected_errors:
                    st.warning(message)

        selected_excel, selected_excel_errors = generate_schedule_workbook(
            [selected_student], schedules, rule_map, gifted_class
        )
        all_excel, all_excel_errors = generate_schedule_workbook(
            edited_students.to_dict("records"), schedules, rule_map, gifted_class
        )
        all_pdf, all_pdf_errors = generate_all_schedule_pdf(
            edited_students.to_dict("records"), schedules, rule_map, gifted_class
        )

        col_download1, col_download2, col_download3 = st.columns(3)
        with col_download1:
            st.download_button(
                "下載這位學生的個人功課表",
                data=selected_excel,
                file_name=f"{target_class}_{int(selected_student['座號']):02d}_{selected_student['姓名']}_個人功課表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_download2:
            st.download_button(
                f"下載{target_class}全班個人功課表",
                data=all_excel,
                file_name=f"{target_class}_全班個人功課表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with col_download3:
            st.download_button(
                f"下載{target_class}全班個人功課表 PDF",
                data=all_pdf,
                file_name=f"{target_class}_全班個人功課表.pdf",
                mime="application/pdf",
            )

        unique_errors = list(dict.fromkeys(all_excel_errors + all_pdf_errors))
        if unique_errors:
            with st.expander(f"全班共有 {len(unique_errors)} 個資料警告", expanded=False):
                for message in unique_errors:
                    st.warning(message)

    except Exception as e:
        st.error(f"個人功課表讀取失敗：{e}")
        st.info("請確認PDF是學校正式課表及原始分組名單；若自動判斷不完整，可在校正表修改。")
else:
    st.info("請先上傳正式課表、數理分組名單及英文分組名單。")
