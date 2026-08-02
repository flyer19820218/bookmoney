import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
import re
import io
import requests
import os

# 🌟 新增：ReportLab PDF 核心繪圖套件
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader

# ==========================================
# 0. 網頁基本設定與自動字型下載 (必須放在最前面)
# ==========================================
st.set_page_config(page_title="校內系統整合平台", layout="centered", page_icon="🏫")

FONT_NAME = "NotoSansTC-Regular.ttf"
FONT_URL = "[https://cdn.jsdelivr.net/gh/themoeway/noto-sans-tc-ttf@master/ttf/NotoSansTC-Regular.ttf](https://cdn.jsdelivr.net/gh/themoeway/noto-sans-tc-ttf@master/ttf/NotoSansTC-Regular.ttf)"

@st.cache_resource
def init_fonts():
    """確保 Streamlit Cloud 上有中文字型可用，避免 PDF 亂碼"""
    if not os.path.exists(FONT_NAME):
        try:
            with st.spinner("正在下載微軟正黑體/Noto中文字型以支援 PDF 產出..."):
                response = requests.get(FONT_URL, headers={'User-Agent': 'Mozilla/5.0'}, timeout=60)
                with open(FONT_NAME, "wb") as f:
                    f.write(response.content)
        except Exception as e:
            st.error(f"字型下載失敗：{e}")
            return False
    try:
        pdfmetrics.registerFont(TTFont('CustomFont', FONT_NAME))
        return True
    except Exception as e:
        st.error(f"字型註冊失敗：{e}")
        return False

HAS_FONT = init_fonts()


# ==========================================
# 頁面 1：書籍費 (您原本的程式碼，一字未改，包裝成函式)
# ==========================================
def page_books():
    st.title("📚 班級各項費用與通知單系統")
    st.markdown("請選擇上方分頁切換您要使用的功能！")

    # 🌟 建立兩個分頁
    tab1, tab2 = st.tabs(["📝 雲端試算表自動版", "🖨️ 懶人 PDF 產生器"])

    # =====================================================================
    # 🌟 分頁 1：原本的雲端自動系統
    # =====================================================================
    with tab1:
        st.header("💡 導師專屬「一鍵排版」全攻略")

        with st.expander("👉 第一步：領取並填寫您的專屬表格 (點我展開)", expanded=True):
            st.markdown("""
            1. 點擊進入 [校內購書公版範本 (短網址：https://reurl.cc/K2LgNe)](https://reurl.cc/K2LgNe)。
            2. 進去後，點選左上角的 **「檔案」 > 「建立副本」**。（一定要建立副本才能編輯喔！）
            3. **不想自己打字？讓 AI 幫你精準整理！**
                手邊只有「紙本估價單」？直接拍照傳給 **Gemini 或 ChatGPT**，並**完整複製貼上以下這段指令給 AI**：
                
                > 你現在是一位專業、細心的「學校資料輸入員」。請將我上傳的書商估價單照片，精準地轉換成表格。
                > 
                > **【表格必須嚴格包含這 4 個欄位】：**
                > 1. **商品名稱**：完整保留書名或材料費名稱。
                > 2. **科目**：請根據書名自動分類為「國、英、數、自、社會、其他」這六類。（注意：歷史、地理、公民請一律歸類為「社會」；聯絡簿、桌墊、材料費請歸類為「其他」）。
                > 3. **分組代號**：若圖片中無特別標示分組，請一律填入數字「1」。如果有分組，請直接填寫代號（如 5A、6B）。
                > 4. **單價**：只填寫純數字，不要加上 $ 或 元。
                > 
                > **【嚴格限制】：** 請勿遺漏項目，不要加入任何問候語或廢話，你的回覆只能有「一個表格」。
                
                接著直接把 AI 做好的表格**全選複製、貼上**到您的 Google 試算表即可！
            """)

        with st.expander("👉 第二步：開啟「共用」權限 (這步沒做會失敗喔！)", expanded=True):
            st.markdown("""
            1. 點擊您 Google 試算表右上角大大的 **「共用」** 按鈕。
            2. 在「一般存取權」下方，將「限制」改為 **「知道連結的人即可檢視」**。
            3. 按下 **「複製連結」**。
            """)

        with st.expander("👉 第三步：貼上網址，領取列印檔", expanded=True):
            st.markdown("""
            1. 將剛剛複製的網址，**貼到下方輸入框**。
            2. 畫面顯示「讀取成功」後，按下 **「🚀 開始產生檔案」**。
            3. 您可以分別下載「給家長的 4格通知單」以及「給導師的 對帳收費總表」。
            """)

        st.warning("⚠️ **列印小叮嚀：** 下載家長通知單後，因各電腦 Excel 預設邊界不同，請在列印時自行點選「預覽列印」微調縮放比例！")
        st.info("💡 **資優生怎麼辦？** 在名單分頁的資優生欄位選「語資」或「數資」，系統會自動扣除。一年級沒分組，代號全填 `1` 就好！")
        st.divider()

        # 原本的核心邏輯函數
        def get_google_sheet_xlsx_url(url):
            try:
                if "docs.google.com" not in url:
                    response = requests.head(url, allow_redirects=True)
                    url = response.url
                    
                pattern = r'https://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)'
                match = re.search(pattern, url)
                if match:
                    sheet_id = match.group(1)
                    return f"[https://docs.google.com/spreadsheets/d/](https://docs.google.com/spreadsheets/d/){sheet_id}/export?format=xlsx"
                return None
            except Exception as e:
                return None

        def should_buy_book(b_subj, b_code, s_gifted, s_eng, s_math, s_sci):
            s_gifted = str(s_gifted).strip()
            b_code = str(b_code).strip()
            
            if s_gifted == "語資" and b_subj in ["國", "英"]: return False
            if s_gifted == "數資" and b_subj in ["數", "自"]: return False
            if b_code in ["1", "全", "", "nan"]: return True
                
            if b_subj == "英" and str(s_eng).strip() in b_code: return True
            if b_subj == "數" and str(s_math).strip() in b_code: return True
            if b_subj == "自" and str(s_sci).strip() in b_code: return True
            return False

        def generate_receipts_excel(df_students, df_books):
            wb = Workbook()
            ws = wb.active
            ws.title = "購書通知單(4張一頁)"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

            ws.column_dimensions['A'].width = 8   ; ws.column_dimensions['B'].width = 38
            ws.column_dimensions['C'].width = 7   ; ws.column_dimensions['D'].width = 2
            ws.column_dimensions['E'].width = 8   ; ws.column_dimensions['F'].width = 38
            ws.column_dimensions['G'].width = 7

            f_title = Font(name="微軟正黑體", size=13, bold=True)
            f_info = Font(name="微軟正黑體", size=12, bold=True)
            f_norm = Font(name="微軟正黑體", size=9)
            f_bold = Font(name="微軟正黑體", size=10, bold=True)
            f_tot = Font(name="微軟正黑體", size=12, bold=True, color="FF0000")
            al_c = Alignment(horizontal="center", vertical="center")
            al_l = Alignment(horizontal="left", vertical="center", wrap_text=True)
            al_r = Alignment(horizontal="right", vertical="center")
            thin = Side(style='thin', color='000000')
            b_all = Border(left=thin, right=thin, top=thin, bottom=thin)

            RECEIPT_ROWS = 12

            for i, row in df_students.iterrows():
                seat, name = row.get("座號", ""), row.get("姓名", "")
                gifted = row.get("資優類別", "")
                eng, math, sci = row.get("英組", "1"), row.get("數組", "1"), row.get("自組", "1")

                student_books = {"國": [], "英": [], "數": [], "自": [], "社會": [], "其他": []}
                for _, book in df_books.iterrows():
                    b_name, b_subj, b_code, b_price = book.get("商品名稱", ""), book.get("科目", ""), book.get("分組代號", "1"), book.get("單價", 0)
                    if b_subj in ["歷", "地", "公", "社會三科"]: b_subj = "社會"
                    if should_buy_book(b_subj, b_code, gifted, eng, math, sci):
                        if b_subj in student_books: student_books[b_subj].append({"name": b_name, "price": b_price})
                        else: student_books["其他"].append({"name": b_name, "price": b_price})

                page = i // 4
                pos = i % 4
                start_row = page * (RECEIPT_ROWS * 2 + 2) + (0 if pos < 2 else RECEIPT_ROWS + 1) + 1
                start_col = 1 if pos % 2 == 0 else 5
                
                ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
                c1 = ws.cell(row=start_row, column=start_col, value="學期購書費通知單")
                c1.font = f_title ; c1.alignment = al_c
                
                ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)
                c2 = ws.cell(row=start_row+1, column=start_col, value=f"座號：{seat}      姓名：{name}")
                c2.font = f_info ; c2.alignment = Alignment(horizontal="left", vertical="center")
                
                for col_offset, text in enumerate(["科目", "購買明細", "小計"]):
                    cell = ws.cell(row=start_row+2, column=start_col+col_offset, value=text)
                    cell.font = f_bold ; cell.alignment = al_c ; cell.border = b_all

                cat_order = ["國", "英", "數", "自", "社會", "其他"]
                curr_row = start_row + 3
                grand_total = 0
                for cat in cat_order:
                    items = student_books[cat]
                    if not items:
                        det_str, subtotal = "免購", 0
                    else:
                        det_str = "、".join([f"{item['name']}(${int(item['price'])})" for item in items])
                        subtotal = sum([item['price'] for item in items])
                    grand_total += subtotal
                    
                    c_cat = ws.cell(row=curr_row, column=start_col, value="社會三科" if cat == "社會" else cat)
                    c_cat.font = f_bold ; c_cat.alignment = al_c ; c_cat.border = b_all
                    c_det = ws.cell(row=curr_row, column=start_col+1, value=det_str)
                    c_det.font = f_norm ; c_det.alignment = al_l ; c_det.border = b_all
                    c_sub = ws.cell(row=curr_row, column=start_col+2, value=subtotal)
                    c_sub.font = f_bold ; c_sub.alignment = al_c ; c_sub.border = b_all
                    ws.row_dimensions[curr_row].height = 28
                    curr_row += 1
                    
                ws.merge_cells(start_row=curr_row, start_column=start_col, end_row=curr_row, end_column=start_col+1)
                c_tot_l = ws.cell(row=curr_row, column=start_col, value="應收總計：")
                c_tot_l.font = f_tot ; c_tot_l.alignment = al_r ; c_tot_l.border = b_all
                ws.cell(row=curr_row, column=start_col+1).border = b_all
                c_tot_v = ws.cell(row=curr_row, column=start_col+2, value=f"{int(grand_total)}")
                c_tot_v.font = f_tot ; c_tot_v.alignment = al_c ; c_tot_v.border = b_all
                ws.row_dimensions[curr_row].height = 25
                
                for r in range(start_row, curr_row + 1):
                    for c in range(start_col, start_col + 3): ws.cell(row=r, column=c).border = b_all

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        def generate_master_excel(df_students, df_books):
            wb = Workbook()
            ws = wb.active
            ws.title = "班級收費總表"
            
            book_rows = []
            for _, book in df_books.iterrows():
                b_name, b_subj, b_code, b_price = book.get("商品名稱", ""), book.get("科目", ""), book.get("分組代號", "1"), book.get("單價", 0)
                b_subj_check = "社會" if b_subj in ["歷", "地", "公", "社會三科"] else b_subj
                
                qty = 0
                for _, s in df_students.iterrows():
                    if should_buy_book(b_subj_check, b_code, s.get("資優類別", ""), s.get("英組", "1"), s.get("數組", "1"), s.get("自組", "1")):
                        qty += 1
                book_rows.append([b_name, b_subj, b_code, qty, b_price])

            student_rows = []
            for _, s in df_students.iterrows():
                seat, name = s.get("座號", ""), s.get("姓名", "")
                gifted = s.get("資優類別", "")
                eng, math, sci = s.get("英組", "1"), s.get("數組", "1"), s.get("自組", "1")
                
                subtotal = 0
                for _, book in df_books.iterrows():
                    b_subj, b_code, b_price = book.get("科目", ""), book.get("分組代號", "1"), book.get("單價", 0)
                    b_subj_check = "社會" if b_subj in ["歷", "地", "公", "社會三科"] else b_subj
                    if should_buy_book(b_subj_check, b_code, gifted, eng, math, sci):
                        subtotal += b_price
                student_rows.append([seat, name, gifted, eng, math, sci, subtotal])

            headers_left = ["商品名稱", "科目", "分組代號", "購買數量", "單價"]
            for col_idx, h in enumerate(headers_left, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

            headers_right = ["座號", "姓名", "資優", "英組", "數組", "自組", "應收總額"]
            for col_idx, h in enumerate(headers_right, 7):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

            thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
                                 top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))

            for r_idx, b_row in enumerate(book_rows, 2):
                for c_idx, val in enumerate(b_row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if c_idx > 1: cell.alignment = Alignment(horizontal="center")

            for r_idx, s_row in enumerate(student_rows, 2):
                for c_idx, val in enumerate(s_row, 7):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if c_idx != 8: cell.alignment = Alignment(horizontal="center")
                    if c_idx == 13: cell.font = Font(bold=True)

            last_row = max(len(book_rows), len(student_rows)) + 2
            ws.cell(row=last_row, column=12, value="班級總計").font = Font(bold=True)
            total_sum = sum([s[-1] for s in student_rows])
            tot_cell = ws.cell(row=last_row, column=13, value=total_sum)
            tot_cell.font = Font(bold=True, color="FF0000"); tot_cell.border = thin_border

            ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 10; ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 3  
            ws.column_dimensions['G'].width = 6 ; ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 8 ; ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 8 ; ws.column_dimensions['L'].width = 8
            ws.column_dimensions['M'].width = 12

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        st.subheader("🛠️ 開始作業")
        default_url = "[https://reurl.cc/K2LgNe](https://reurl.cc/K2LgNe)"
        sheet_url = st.text_input("👇 請在下方輸入您的試算表網址：", value=default_url)

        if sheet_url:
            xlsx_url = get_google_sheet_xlsx_url(sheet_url)
            if xlsx_url:
                try:
                    with st.spinner("正在連線至 Google Sheets 讀取資料..."):
                        df_students = pd.read_excel(xlsx_url, sheet_name=0).fillna("")
                        df_books = pd.read_excel(xlsx_url, sheet_name=1).fillna("")
                    
                    st.success(f"✅ 讀取成功！共載入 {len(df_students)} 位學生、{len(df_books)} 筆書目報價。")
                    
                    if st.button("🚀 開始產生檔案", type="primary"):
                        with st.spinner("系統正在計算與排版中..."):
                            receipts_data = generate_receipts_excel(df_students, df_books)
                            master_data = generate_master_excel(df_students, df_books)
                            
                        st.balloons()
                        col1, col2 = st.columns(2)
                        with col1:
                            st.download_button(label="📥 下載【家長通知單】(A4列印版)", data=receipts_data, file_name="家長購書通知單.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                        with col2:
                            st.download_button(label="📥 下載【導師對帳總表】(收費明細)", data=master_data, file_name="導師收費對帳總表.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error("❌ 讀取失敗！請確認網址與權限。")

    # =====================================================================
    # 🌟 分頁 2：班級購書與費用對帳系統
    # =====================================================================
    with tab2:
        st.subheader("🖨️ 班級購書與費用對帳系統")
        
        # 🌟 將使用說明直接顯示在網頁最上方
        st.info("""
        **【系統使用說明】**
        1. 請至「台中市教育雲端系統」下載 **班級名條** (若為有分組之年級，請一併下載英文/數學分組名條)。
        2. 請向各家書商索取本次的 **報價單 CSV 或 Excel 檔**。
        3. 依序上傳下方檔案，若有其他自訂收費項目，可於步驟二表格手動新增。
        *(註：一年級免分組，僅需上傳「班級名單」與「書商報價單」即可，系統會自動產出全班明細)*
        """)

        st.markdown("#### 📁 步驟一：上傳名單與報價單")

        col_a, col_b = st.columns(2)
        with col_a:
            file_class = st.file_uploader("1. 班級總名單 (Excel 檔)", type=["xlsx", "xls"])
            file_books_list = st.file_uploader("2. 書商報價單 (可框選多份 CSV/Excel)", type=["csv", "xlsx", "xls"], accept_multiple_files=True)
        with col_b:
            file_eng   = st.file_uploader("3. 英文分組名單 (可選)", type=["csv", "xlsx", "xls"])
            file_math  = st.file_uploader("4. 數學/自然分組名單 (可選)", type=["csv", "xlsx", "xls"])
            st.caption("💡 提示：系統若讀到 9 或 3 開頭的三年級名單，將自動連動數/自為同組。")

        st.markdown("#### ➕ 步驟二：新增其他收費 (選填)")
        st.write("若無其他收費請留空。填寫「分組代號」系統會自動結算人數 (全班收取請填 `1`)。")
        
        if 'custom_fees' not in st.session_state:
            st.session_state.custom_fees = pd.DataFrame([
                {"商品名稱": "", "科目": "其他", "分組代號": "1", "單價": 0}
            ])
            
        edited_custom_fees = st.data_editor(
            st.session_state.custom_fees,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "商品名稱": st.column_config.TextColumn("商品名稱 (必填)"),
                "科目": st.column_config.SelectboxColumn("科目分類", options=["國", "英", "數", "自", "社會", "其他"]),
                "分組代號": st.column_config.TextColumn("分組代號 (全班請填 1)"),
                "單價": st.column_config.NumberColumn("每人單價 (元)", min_value=0, step=1),
            }
        )

        if 'pdf_output' not in st.session_state: st.session_state.pdf_output = None
        if 'excel_output' not in st.session_state: st.session_state.excel_output = None

        def find_column(df, keywords, default_name):
            for col in df.columns:
                if any(kw in str(col) for kw in keywords): return col
            return None
            
        def guess_subject(name, code=""):
            name = str(name).replace("國中", "").replace("國小", "")
            if any(k in name for k in ["英", "文法", "單字", "聽力", "ABC", "abc"]): return "英"
            if any(k in name for k in ["數", "幾何", "代數", "算"]): return "數"
            if any(k in name for k in ["自", "理化", "生物", "地科", "科學"]): return "自"
            if any(k in name for k in ["歷", "地", "公", "社會"]): return "社會"
            if any(k in name for k in ["國", "文", "閱讀"]): return "國"
            
            code_str = str(code).replace(" ", "").upper()
            if "英" in code_str: return "英"
            if "數" in code_str: return "數"
            if "自" in code_str: return "自"
            return "其他"

        def standardize_book_code(val):
            val = str(val).strip()
            if val in ["1", "全", "", "nan", "None", "無"]: return "1"
            match = re.search(r'(\d+).*?([A-Za-z])', val)
            if match:
                cls_num = int(match.group(1))
                if cls_num >= 100: cls_num = cls_num % 100 
                return f"{cls_num}{match.group(2).upper()}"
            match_alpha = re.search(r'[A-Za-z]', val)
            if match_alpha: return match_alpha.group(0).upper()
            return val

        # 🌟 一字不漏還原：完美處理 Excel 合併儲存格的繼承邏輯
        def parse_horizontal_group_file(uploaded_file, subj_hint=""):
            if uploaded_file.name.endswith('.csv'): df_grp = pd.read_csv(uploaded_file, header=None).fillna("")
            else: df_grp = pd.read_excel(uploaded_file, header=None).fillna("")
            header_idx = -1
            for idx, row in df_grp.iterrows():
                if any(isinstance(v, str) and "姓名" in v for v in row.values):
                    header_idx = idx ; break
            mapping = {}
            if header_idx != -1:
                col_to_group = {}
                current_group = "無"
                current_grade = 0
                for col_idx in range(df_grp.shape[1]):
                    for r in range(0, header_idx):
                        val = str(df_grp.iloc[r, col_idx]).strip()
                        if val and val != "nan":
                            match_num_alpha = re.search(r'(\d+).*?([A-Za-z])', val)
                            if match_num_alpha:
                                cls_num = int(match_num_alpha.group(1))
                                current_grade = cls_num // 100 
                                if cls_num >= 100: cls_num = cls_num % 100 
                                current_group = f"{subj_hint}{cls_num}{match_num_alpha.group(2).upper()}"
                            else:
                                match_alpha = re.search(r'[A-Za-z]', val)
                                if match_alpha: current_group = f"{subj_hint}{match_alpha.group(0).upper()}"
                                elif len(val) <= 4: current_group = val
                    col_to_group[col_idx] = (current_group, current_grade)

                for col_idx in range(df_grp.shape[1]):
                    if "姓名" in str(df_grp.iloc[header_idx, col_idx]):
                        group_info = col_to_group.get(col_idx, ("無", 0))
                        names = df_grp.iloc[header_idx+1:, col_idx].astype(str).str.strip()
                        for name in names:
                            clean_name = name.replace(" ", "")
                            if clean_name and clean_name != "nan":
                                mapping[clean_name] = group_info
            return mapping

        def check_group_match(s_grp, b_code):
            if s_grp in ["1", "無", "免", ""]: return False
            b_clean = str(b_code).strip().upper()
            s_clean = str(s_grp).strip().upper()
            if s_clean == b_clean: return True
            
            s_match = re.search(r'(\d*)([A-Z]+)', s_clean)
            b_match = re.search(r'(\d*)([A-Z]+)', b_clean)
            if s_match and b_match:
                s_num, s_letter = s_match.groups()
                b_num, b_letter = b_match.groups()
                if s_letter != b_letter: return False
                if b_num and s_num and b_num != s_num: return False
                return True
            return False

        def safe_psychic_correction(df_b, df_s):
            counts = {
                '英': df_s[~df_s['英組'].isin(['無', '免', '1', ''])]['英組'].value_counts().to_dict(),
                '數': df_s[~df_s['數組'].isin(['無', '免', '1', ''])]['數組'].value_counts().to_dict(),
                '自': df_s[~df_s['自組'].isin(['無', '免', '1', ''])]['自組'].value_counts().to_dict()
            }
            
            for idx, row in df_b.iterrows():
                b_code = str(row['code']).strip().upper()
                b_subj = str(row['subj'])
                b_qty = row['qty']
                
                if b_qty > 0 and re.fullmatch(r'[A-Z]', b_code) and b_subj in counts:
                    subj_counts = counts[b_subj]
                    possible_fixes = []
                    for grp_name, grp_count in subj_counts.items():
                        if grp_count == b_qty and b_code in grp_name:
                            possible_fixes.append(grp_name)
                            
                    if len(possible_fixes) == 1:
                        df_b.at[idx, 'code'] = possible_fixes[0]
                        
            return df_b

        def is_book_for_student(b_subj, b_code, s_eng, s_math, s_sci, s_gifted):
            if s_gifted == "語資" and b_subj in ["國", "英"]: return False
            if s_gifted == "數資" and b_subj in ["數", "自"]: return False
            if "語資" in s_gifted and "數資" in s_gifted and b_subj in ["國", "英", "數", "自"]: return False
            
            if b_code == "1": return True
            
            if b_subj == "英": return check_group_match(s_eng, b_code)
            if b_subj == "數": return check_group_match(s_math, b_code)
            if b_subj == "自": return check_group_match(s_sci, b_code)
            if b_subj not in ["英", "數", "自"]: 
                return check_group_match(s_eng, b_code) or check_group_match(s_math, b_code)
            return False

        def generate_smart_pdf(df_students, df_books_clean):
            pdf_io = io.BytesIO()
            c = canvas.Canvas(pdf_io, pagesize=A4)
            width, height = A4
            pdf_font = 'CustomFont' if HAS_FONT else 'Helvetica'
            
            for _, student in df_students.iterrows():
                seat = str(student.get("座號", "")).split('.')[0]
                name = str(student.get("姓名", "")).strip()
                s_eng = str(student.get("英組", "無")).strip()
                s_math = str(student.get("數組", "無")).strip()
                s_sci = str(student.get("自組", "無")).strip()
                s_gifted = str(student.get("資優類別", "無")).strip()

                personal_list = []
                total_amount = 0
                for _, b in df_books_clean.iterrows():
                    if is_book_for_student(b['subj'], b['code'], s_eng, s_math, s_sci, s_gifted):
                        personal_list.append((b['name'], b['price']))
                        total_amount += b['price']

                c.setFont(pdf_font, 24)
                c.drawCentredString(width/2, height - 80, "學 期 各 項 費 用 通 知 單")
                s_eng_display = s_eng if s_eng not in ["1", "無", "免"] else "無"
                s_math_display = s_math if s_math not in ["1", "無", "免"] else "無"
                s_gifted_display = s_gifted if s_gifted not in ["1", "無", "免"] else "無"
                
                c.setFont(pdf_font, 14)
                c.drawString(60, height - 130, f"座號：{seat}        姓名：{name}")
                c.drawRightString(width - 60, height - 130, f"狀態：英({s_eng_display}) 數({s_math_display}) 資優({s_gifted_display})")
                c.setStrokeColorRGB(0, 0, 0)
                c.setLineWidth(2)
                c.line(60, height - 145, width - 60, height - 145)
                
                c.setFont(pdf_font, 12)
                c.drawString(70, height - 170, "項目名稱")
                c.drawRightString(width - 70, height - 170, "金額 (元)")
                c.setLineWidth(0.5)
                c.line(60, height - 180, width - 60, height - 180)
                
                start_y = height - 205
                end_y = 160 
                available_space = start_y - end_y
                item_count = len(personal_list)
                
                if item_count == 0:
                    c.setFont(pdf_font, 12)
                    c.drawString(70, start_y, "（本學期無收費項目，免繳費）")
                    y_pos = start_y - 25
                    line_height = 20
                else:
                    default_line_height = 22 
                    if item_count * default_line_height > available_space:
                        line_height = available_space / item_count
                        dynamic_font_size = max(9, int(line_height * 0.6)) 
                    else:
                        line_height = default_line_height
                        dynamic_font_size = 12
                    
                    c.setFont(pdf_font, dynamic_font_size)
                    y_pos = start_y
                    for b_name, price in personal_list:
                        c.drawString(70, y_pos, b_name)
                        c.drawRightString(width - 70, y_pos, f"$ {int(price)}")
                        y_pos -= line_height
                
                c.setLineWidth(1)
                c.line(60, y_pos + (line_height * 0.6), width - 60, y_pos + (line_height * 0.6))
                c.setFont(pdf_font, 16)
                c.drawString(70, y_pos - 20, "應繳總計金額：")
                c.setFillColorRGB(0.8, 0, 0)
                c.drawRightString(width - 70, y_pos - 20, f"$ {int(total_amount)} 元")
                c.setFillColorRGB(0, 0, 0)
                
                box_y = 55
                c.setFont(pdf_font, 12)
                c.drawString(60, box_y + 55, "【家長簽章回條】")
                c.setFont(pdf_font, 11)
                c.drawString(60, box_y + 25, f"本人已確認上述 座號 {seat} {name} 之明細與金額無誤。")
                c.drawString(width - 150, box_y + 25, "家長簽名：")
                c.showPage()
            c.save()
            pdf_io.seek(0)
            return pdf_io

        def fill_sheet_data(ws, sheet_title, df_students, df_books):
            ws.title = sheet_title
            book_rows = []
            for _, book in df_books.iterrows():
                b_name, b_subj, b_price = book['name'], book['subj'], book['price']
                b_code_str = str(book['code']).strip()
                qty = 0
                for _, s in df_students.iterrows():
                    s_eng, s_math, s_sci = str(s.get("英組","無")), str(s.get("數組","無")), str(s.get("自組","無"))
                    s_gifted = str(s.get("資優類別", "無"))
                    if is_book_for_student(b_subj, b_code_str, s_eng, s_math, s_sci, s_gifted): qty += 1
                book_rows.append([b_name, b_subj, b_code_str, qty, b_price])

            student_rows = []
            for _, s in df_students.iterrows():
                seat, name = str(s.get("座號", "")).split('.')[0], s.get("姓名", "")
                gifted = str(s.get("資優類別", "無"))
                s_eng, s_math, s_sci = str(s.get("英組","無")), str(s.get("數組","無")), str(s.get("自組","無"))
                subtotal = 0
                for _, book in df_books.iterrows():
                    if is_book_for_student(book['subj'], book['code'], s_eng, s_math, s_sci, gifted):
                        subtotal += book['price']
                student_rows.append([seat, name, gifted, s_eng, s_math, s_sci, subtotal])

            headers_left = ["項目名稱", "科目", "分組代號", "結算數量", "單價"]
            for col_idx, h in enumerate(headers_left, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

            headers_right = ["座號", "姓名", "資優", "英組", "數組", "自組", "應收總額"]
            for col_idx, h in enumerate(headers_right, 7):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")

            thin_border = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
            for r_idx, b_row in enumerate(book_rows, 2):
                for c_idx, val in enumerate(b_row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if c_idx > 1: cell.alignment = Alignment(horizontal="center")

            for r_idx, s_row in enumerate(student_rows, 2):
                for c_idx, val in enumerate(s_row, 7):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = thin_border
                    if c_idx != 8: cell.alignment = Alignment(horizontal="center")
                    if c_idx == 13: cell.font = Font(bold=True)

            last_row = max(len(book_rows), len(student_rows)) + 2
            ws.cell(row=last_row, column=12, value="單頁總計").font = Font(bold=True)
            total_sum = sum([s[-1] for s in student_rows])
            tot_cell = ws.cell(row=last_row, column=13, value=total_sum)
            tot_cell.font = Font(bold=True, color="FF0000"); tot_cell.border = thin_border
            
            ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 10; ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 3  
            ws.column_dimensions['G'].width = 6 ; ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 8 ; ws.column_dimensions['J'].width = 8
            ws.column_dimensions['K'].width = 8 ; ws.column_dimensions['L'].width = 8
            ws.column_dimensions['M'].width = 12

        def generate_excel_master_dynamic(df_students, df_books_clean):
            wb = Workbook()
            ws1 = wb.active
            fill_sheet_data(ws1, "班級收費總表(全)", df_students, df_books_clean)
            
            unique_publishers = df_books_clean['publisher'].unique()
            for pub in unique_publishers:
                df_pub = df_books_clean[df_books_clean['publisher'] == pub]
                ws_pub = wb.create_sheet()
                safe_pub_name = str(pub).replace("/", "").replace("\\", "")[:12]
                fill_sheet_data(ws_pub, f"{safe_pub_name}對帳表", df_students, df_pub)
                
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # 🌟 執行區塊
        if file_class and file_books_list and len(file_books_list) > 0:
            try:
                df_temp = pd.read_excel(file_class, header=None).fillna("")
                header_idx = 0
                for idx, row in df_temp.iterrows():
                    if any("姓名" in str(v) for v in row.values):
                        header_idx = idx ; break
                df_s = pd.read_excel(file_class, skiprows=header_idx).fillna("")
                
                c_name = find_column(df_s, ["姓名", "名稱", "學生"], "姓名")
                if c_name: df_s = df_s[df_s[c_name].astype(str).str.strip() != ""] 
                
                c_seat = find_column(df_s, ["座號", "號碼", "序號"], "座號")
                if c_seat and c_seat != "座號": df_s = df_s.rename(columns={c_seat: "座號"})
                if c_name and c_name != "姓名": df_s = df_s.rename(columns={c_name: "姓名"})
                
                for col in ["英組", "數組", "自組", "資優類別"]:
                    if col not in df_s.columns: df_s[col] = "無"

                eng_map = parse_horizontal_group_file(file_eng, "英") if file_eng else {}
                math_map = parse_horizontal_group_file(file_math, "數") if file_math else {}

                for idx, row in df_s.iterrows():
                    clean_name = str(row["姓名"]).replace(" ", "").strip()
                    
                    if file_eng:
                        if clean_name in eng_map: df_s.at[idx, "英組"] = eng_map[clean_name][0]
                        else:
                            df_s.at[idx, "資優類別"] = "語資"
                            df_s.at[idx, "英組"] = "免"
                            
                    if file_math:
                        if clean_name in math_map:
                            grp_code, grade = math_map[clean_name]
                            df_s.at[idx, "數組"] = grp_code
                            if grade in [9, 3]: 
                                df_s.at[idx, "自組"] = grp_code  
                            else:
                                df_s.at[idx, "自組"] = "免"
                        else:
                            current_gifted = str(df_s.at[idx, "資優類別"])
                            if current_gifted in ["無", "", "1"]: df_s.at[idx, "資優類別"] = "數資"
                            elif current_gifted == "語資": df_s.at[idx, "資優類別"] = "語資/數資"
                            df_s.at[idx, "數組"] = "免"
                            df_s.at[idx, "自組"] = "免"

                with st.expander("👀 預覽：學生名單與分組狀態"):
                    st.write("請確認名單與分組判定是否正確 (三年級數/自已設定為自動連動)：")
                    st.dataframe(df_s[["座號", "姓名", "英組", "數組", "自組", "資優類別"]])

                all_books_clean_list = []
                
                for fb in file_books_list:
                    publisher_name = fb.name.split('.')[0]
                    extracted_books = []
                    header_skip = 0
                    keywords_header = ["品名", "商品", "名稱", "單價", "價格", "金額", "數量", "件數"]
                    
                    if fb.name.endswith('.csv'):
                        bytes_data = fb.read()
                        lines = []
                        used_enc = 'utf-8'
                        for enc in ['utf-8', 'big5', 'cp950', 'utf-8-sig']:
                            try:
                                lines = bytes_data.decode(enc).splitlines()
                                used_enc = enc ; break
                            except: continue
                        
                        for idx, line in enumerate(lines):
                            if sum(1 for kw in keywords_header if kw in line) >= 2:
                                header_skip = idx ; break
                                
                        if header_skip > 0 and len(lines) > 0:
                            clean_first_line = lines[0].replace(',', '').strip()
                            if clean_first_line:
                                publisher_name = clean_first_line.replace("估價單", "").replace("報價單", "").strip()
                                
                        fb.seek(0)
                        try: df_b = pd.read_csv(fb, skiprows=header_skip, encoding=used_enc).fillna("")
                        except: 
                            fb.seek(0)
                            df_b = pd.read_csv(fb, skiprows=header_skip, on_bad_lines='skip').fillna("")
                    
                    else:
                        df_temp2 = pd.read_excel(fb, header=None).fillna("")
                        for idx, row in df_temp2.iterrows():
                            row_str = "".join([str(val) for val in row.values])
                            if sum(1 for kw in keywords_header if kw in row_str) >= 2:
                                header_skip = idx ; break
                                
                        if header_skip > 0:
                            first_row_str = "".join([str(val) for val in df_temp2.iloc[0].values if str(val).strip() and str(val) != "nan"])
                            if first_row_str:
                                publisher_name = first_row_str.replace("估價單", "").replace("報價單", "").strip()
                                
                        fb.seek(0)
                        df_b = pd.read_excel(fb, skiprows=header_skip).fillna("")

                    b_col_name = find_column(df_b, ["品名", "商品", "名稱", "書籍"], "商品名稱")
                    b_col_price = find_column(df_b, ["單價", "價格", "金額"], "單價")
                    b_col_code = find_column(df_b, ["附記", "備註", "分組", "代號"], "分組代號")
                    b_col_qty = find_column(df_b, ["數量", "量", "件數"], "數量")
                    b_col_subj = find_column(df_b, ["科目", "類別"], "科目")
                    
                    df_b['parsed_price'] = pd.to_numeric(df_b[b_col_price], errors='coerce')
                    
                    for i in range(len(df_b)):
                        price_val = df_b['parsed_price'].iloc[i]
                        if pd.notna(price_val) and price_val > 0: 
                            raw_name = str(df_b[b_col_name].iloc[i])
                            raw_code = standardize_book_code(df_b[b_col_code].iloc[i]) if b_col_code else "1"
                            raw_qty = df_b[b_col_qty].iloc[i] if b_col_qty else 0
                            qty_val = pd.to_numeric(raw_qty, errors='coerce')
                            qty_val = int(qty_val) if pd.notna(qty_val) else 0
                            
                            subj_val = str(df_b[b_col_subj].iloc[i]) if b_col_subj else guess_subject(raw_name, raw_code)

                            extracted_books.append({
                                'name': raw_name,
                                'price': int(price_val),
                                'code': raw_code,
                                'qty': qty_val,
                                'subj': subj_val,
                                'publisher': publisher_name
                            })

                    if extracted_books:
                        df_temp_clean = pd.DataFrame(extracted_books)
                        all_books_clean_list.append(df_temp_clean)
                
                custom_books = []
                for _, row in edited_custom_fees.iterrows():
                    name_val = str(row.get("商品名稱", "")).strip()
                    price_val = pd.to_numeric(row.get("單價", 0), errors='coerce')
                    if name_val and name_val != "nan" and pd.notna(price_val) and price_val > 0:
                        c_subj = str(row.get("科目", "其他"))
                        c_code = standardize_book_code(row.get("分組代號", "1"))
                        custom_books.append({
                            'name': name_val,
                            'price': int(price_val),
                            'code': c_code,
                            'qty': 0,
                            'subj': c_subj,
                            'publisher': "新增自訂項目"
                        })
                if custom_books:
                    all_books_clean_list.append(pd.DataFrame(custom_books))
                    
                df_books_clean = pd.concat(all_books_clean_list, ignore_index=True) if all_books_clean_list else pd.DataFrame()

                if not df_books_clean.empty:
                    df_books_clean = safe_psychic_correction(df_books_clean, df_s)
                    df_books_clean['subj'] = df_books_clean['subj'].apply(lambda x: "社會" if x in ["歷", "地", "公", "歷史", "地理", "公民"] else x)

                with st.expander("👀 預覽：書目與收費清單"):
                    st.write("請確認系統讀取的項目與單價是否正確：")
                    if not df_books_clean.empty:
                        st.dataframe(df_books_clean)
                    else:
                        st.warning("⚠️ 尚未成功讀取任何書目。")

                st.divider()
                st.markdown("#### 🚀 步驟三：產生報表")

                if not df_books_clean.empty and st.button("確認無誤，開始產出檔案", type="primary"):
                    with st.spinner("正在產生 PDF 與 Excel 檔案中..."):
                        st.session_state.pdf_output = generate_smart_pdf(df_s, df_books_clean)
                        st.session_state.excel_output = generate_excel_master_dynamic(df_s, df_books_clean)
                        st.success("✅ 檔案產生完成！請點擊下方按鈕下載。")

                if st.session_state.pdf_output and st.session_state.excel_output:
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 下載【家長通知單】(PDF 檔)", 
                            data=st.session_state.pdf_output, 
                            file_name="全班通知單_自動對帳版.pdf", 
                            mime="application/pdf"
                        )
                    with col2:
                        st.download_button(
                            label="📥 下載【導師對帳總表】(Excel 檔)", 
                            data=st.session_state.excel_output, 
                            file_name="導師總表_收費明細版.xlsx", 
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            
            except Exception as e:
                st.error(f"系統讀取發生錯誤：{e}")
                st.info("提示：請檢查上傳的檔案格式是否正確。")
        else:
            st.info("💡 請先依序上傳名單與報價單，預覽畫面將自動顯示。")


# ==========================================
# 頁面 2：成績單產生器 (請將您原本的成績單程式碼貼在這邊)
# ==========================================
def page_grades():
    st.title("📝 成績單產生器")
    st.write("---")
    st.info("👋 這裡已經為你準備好「成績單」專屬頁面了！")
    st.write("請將您之前寫好的成績單處理程式碼 (UI 與運算邏輯)，直接貼在 `def page_grades():` 這個函式裡面即可。")


# ==========================================
# 頁面 3：退休金試算
# ==========================================
def page_pension():
    st.title("💰 教師退休金（月退俸）試算")
    st.info("💡 財務評估原則：本試算依據現行公式得出，僅提供最保守之基準參考，不計入未來法規變動或通膨等外部因素。")

    # 薪級字典
    SALARY_GRADES = {
        680: 57220, 650: 55690, 625: 54160, 600: 52630, 575: 51100, 550: 49560, 
        525: 48030, 500: 46500, 475: 44970, 450: 41900, 430: 40760, 410: 39610, 
        390: 38460, 370: 37310, 350: 36160, 330: 35010, 310: 33860, 290: 32710, 
        275: 31560, 260: 30410, 245: 29270, 230: 28120, 220: 27350, 210: 26580, 
        200: 25820, 190: 25050
    }

    def calculate_replacement_ratio(years):
        if years < 15: return 0.0
        elif years == 15: return 0.39
        elif 15 < years <= 35: return 0.39 + (years - 15) * 0.015
        elif 35 < years <= 40: return 0.69 + (years - 35) * 0.005
        else: return 0.715

    st.markdown("### 1. 基本資料輸入")
    col1, col2 = st.columns(2)
    
    with col1:
        years = st.number_input("教學年資（年）", min_value=1, max_value=50, value=25, step=1, key="pension_years")
    
    with col2:
        input_method = st.radio("最後15年平均本俸設定方式", ["從薪級表選擇 (自動帶入)", "自行手動輸入"], key="pension_method")

    if input_method == "從薪級表選擇 (自動帶入)":
        selected_grade = st.selectbox(
            "請選擇薪級（最低從 190 起算）", 
            options=list(SALARY_GRADES.keys()), 
            index=list(SALARY_GRADES.keys()).index(600)
        )
        avg_base_salary = SALARY_GRADES[selected_grade]
        st.write(f"**對應薪額：** {avg_base_salary:,} 元")
    else:
        avg_base_salary = st.number_input(
            "請輸入最後15年平均本俸（元）", 
            min_value=25050, 
            value=52630, 
            step=1000,
            key="pension_salary"
        )

    st.divider()
    st.markdown("### 2. 試算結果")

    if years < 15:
        st.warning("⚠️ 依據規定，年資需滿 15 年始得請領月退俸。")
    else:
        ratio = calculate_replacement_ratio(years)
        monthly_pension = avg_base_salary * 2 * ratio

        st.markdown(f"""
        - **所得替代率：** {ratio * 100:.1f} %
        - **計算公式：** {avg_base_salary:,} × 2 × {ratio:.3f}
        """)
        
        st.metric(label="預估月退俸金額", value=f"NT$ {int(monthly_pension):,}")


# ==========================================
# 主程式導覽設定 (負責渲染左側選單與切換)
# ==========================================
def main():
    st.sidebar.title("🏫 校內系統整合平台")
    st.sidebar.write("請從下方選擇您要使用的工具：")
    
    # 建立左側選單按鈕
    page = st.sidebar.radio(
        "功能選單",
        ["書籍費", "成績單產生器", "退休金"]
    )
    
    st.sidebar.divider()
    st.sidebar.caption("系統版本：v2.0")

    # 根據選擇切換執行對應的函式
    if page == "書籍費":
        page_books()
    elif page == "成績單產生器":
        page_grades()
    elif page == "退休金":
        page_pension()

if __name__ == "__main__":
    main()
